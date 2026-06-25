import csv
from pathlib import Path
from typing import Iterable, Sequence

import xlrd
from openpyxl import load_workbook


def _clean_cell(value: object) -> str:
    return str(value).strip()


def _compact_row(values: Iterable[object]) -> list[str]:
    return [
        _clean_cell(value)
        for value in values
        if value is not None and _clean_cell(value)
    ]


def _format_row(values: Iterable[object]) -> str:
    cells = [
        _clean_cell(value)
        for value in values
        if value is not None and _clean_cell(value)
    ]

    return " | ".join(cells)


def _looks_like_header(row: Sequence[str]) -> bool:
    if not row:
        return False

    numeric_like = 0

    for cell in row:
        normalized = cell.replace(".", "").replace(",", "").replace("/", "")

        if normalized.isdigit():
            numeric_like += 1

    return numeric_like == 0


def _default_headers(width: int) -> list[str]:
    return [f"Kolom {index}" for index in range(1, width + 1)]


def _format_record(
    values: Sequence[object],
    headers: Sequence[str],
    row_number: int,
    sheet_name: str | None = None,
) -> str:
    pairs = []

    for index, value in enumerate(values):
        cell = _clean_cell(value)

        if not cell:
            continue

        header = headers[index] if index < len(headers) else f"Kolom {index + 1}"
        pairs.append(f"{header}: {cell}")

    if not pairs:
        return ""

    prefix = f"Sheet: {sheet_name}; " if sheet_name else ""
    return f"{prefix}Baris {row_number}; " + "; ".join(pairs)


def _build_record_pages(
    raw_rows: list[Sequence[object]],
    sheet_index: int,
    sheet_name: str | None,
) -> list[dict]:
    compact_rows = [_compact_row(row) for row in raw_rows]
    first_row_index = next(
        (
            index for index, row in enumerate(compact_rows)
            if row
        ),
        None,
    )

    if first_row_index is None:
        return []

    first_row = compact_rows[first_row_index]
    has_header = _looks_like_header(first_row)
    headers = first_row if has_header else _default_headers(len(first_row))
    data_start_index = first_row_index + 1 if has_header else first_row_index

    pages = []

    for row_index in range(data_start_index, len(raw_rows)):
        row = raw_rows[row_index]
        text = _format_record(
            values=row,
            headers=headers,
            row_number=row_index + 1,
            sheet_name=sheet_name,
        )

        if not text:
            continue

        pages.append({
            "page_number": sheet_index,
            "text": text,
            "extraction_method": "spreadsheet",
            "sheet_name": sheet_name,
            "row_number": row_index + 1,
        })

    return pages


def _extract_xlsx_pages(file_path: Path) -> list[dict]:
    workbook = load_workbook(
        filename=file_path,
        read_only=True,
        data_only=True,
    )

    pages = []

    try:
        for sheet_index, worksheet in enumerate(workbook.worksheets, start=1):
            rows = list(worksheet.iter_rows(values_only=True))
            pages.extend(_build_record_pages(
                raw_rows=rows,
                sheet_index=sheet_index,
                sheet_name=worksheet.title,
            ))
    finally:
        workbook.close()

    return pages


def _extract_csv_pages(file_path: Path) -> list[dict]:
    with file_path.open("r", encoding="utf-8-sig", newline="") as csv_file:
        reader = csv.reader(csv_file)
        rows = list(reader)

    return _build_record_pages(
        raw_rows=rows,
        sheet_index=1,
        sheet_name=None,
    )


def _extract_xls_pages(file_path: Path) -> list[dict]:
    workbook = xlrd.open_workbook(file_path)

    pages = []

    for sheet_index in range(workbook.nsheets):
        worksheet = workbook.sheet_by_index(sheet_index)
        rows = [
            worksheet.row_values(row_index)
            for row_index in range(worksheet.nrows)
        ]
        pages.extend(_build_record_pages(
            raw_rows=rows,
            sheet_index=sheet_index + 1,
            sheet_name=worksheet.name,
        ))

    return pages


def extract_spreadsheet_pages(file_path: Path, extension: str) -> list[dict]:
    if extension == ".xlsx":
        return _extract_xlsx_pages(file_path)

    if extension == ".xls":
        return _extract_xls_pages(file_path)

    if extension == ".csv":
        return _extract_csv_pages(file_path)

    raise ValueError(f"Ekstensi spreadsheet tidak didukung: {extension}")
